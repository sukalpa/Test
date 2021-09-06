import openpyxl

def getRowCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_row)

def getColumnCount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_column)

def readdata(file,sheetname,rowno,colno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rowno,column=colno).value

def writedata(file,sheetname,rowno,colno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rowno,column=colno).value=data
    workbook.save(file)

def readfromexcel(file,sheetname):
    # workbook object is created
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    m_row = sheet.max_row
    # Loop will print all values
    # of first column
    for i in range(2, m_row + 1):
        cell_obj = sheet.cell(row=i, column=2)
        return cell_obj.value

